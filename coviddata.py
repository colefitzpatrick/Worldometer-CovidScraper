from selenium import webdriver
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import os
import time
import openpyxl

os.chdir('c:\\Python\\colefitzpatrick_python')

wb_write = openpyxl.load_workbook('covid.xlsx')
ws_write = wb_write["Sheet1"]

url = "https://www.worldometers.info/coronavirus/country/us/"
driver = webdriver.Firefox()
driver.implicitly_wait(1)
driver.get(url)

d_full = datetime.today() - timedelta(days=1)
dm = d_full.strftime("%m")
dd = d_full.strftime("%d")
dy = d_full.strftime("%Y")
d = dm + "/" + dd + "/" + dy

print(d)

page_data=BeautifulSoup(driver.page_source, 'lxml')

table = page_data.find("table", {"id": "usa_table_countries_yesterday"})

writerow = ws_write.max_row + 1
print(ws_write.max_row)


def table_to_excel( str ):
    global writerow
    for tr in table.findAll("tr", {"class": str}):
        row = [d]
        writecol = 2
        
        for td in tr.findAll("td"):
            if len(row) == 1:
                row.append(td.text.strip())
            else:
                nocomma = td.text.replace(',', '')
                noplus = nocomma.replace('+', '')
                nospace = noplus.replace(' ', '')
                row.append(nospace)
        ws_write.cell(row=writerow, column=1).value = d        
        for number in range(1,11):
            if number != 1 and len(row[number]) > 0:
                ws_write.cell(row=writerow, column=writecol).value = int(row[number])
            else:
                ws_write.cell(row=writerow, column=writecol).value = row[number]
            writecol += 1
        writerow += 1
    return

table_to_excel("even")
table_to_excel("odd")

wb_write.save('covid.xlsx')
